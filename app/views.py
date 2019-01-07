from django.shortcuts import render
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import render
from django.core.files.storage import default_storage
import openpyxl
import googlemaps
import uuid
from geocoder.settings import MEDIA_ROOT, MEDIA_URL, GOOGLE_API_KEY
import logging
import datetime
LOG_FILENAME = 'error.log'
logging.basicConfig(filename=LOG_FILENAME,level=logging.DEBUG)

def main(request):
    if request.method == "GET":
        return render(request, 'main.html', {})
    else:
        try:
            addressfile = request.FILES["file"]
            book = openpyxl.load_workbook(addressfile)
            if len(book.sheetnames) < 1:
                return JsonResponse({"Success":False, "Error": "No sheets found"})
            sheetname = book.sheetnames[0]
            sheet = book[sheetname]
            data = []
            row_count = sheet.max_row
            for index in range(1, row_count + 1):
                data.append(sheet.cell(row = index, column = 1).value)
            return JsonResponse(createGeoLocationFile(data, populateGeocodes(data), sheetname))
        except Exception as exception:
            logging.error(str(datetime.datetime.now()) + ":" + exception.__str__())
            return JsonResponse({'success': False, 'error_text': "An error has occurred while processing the request."})

def populateGeocodes(data):
    client = googlemaps.Client(key=GOOGLE_API_KEY)
    results = client.geocode(data)#['1600 Amphitheatre Parkway, Mountain View, CA', 'Ittina Neela, Andapura, Bangalore'])
    result_data = list()
    for result in results:
        result_data.append(result["geometry"]["location"])
    return result_data

def createGeoLocationFile(input, results, title):
    book = openpyxl.Workbook()  
    sheet = book.active 
    sheet.title = title 
    count = len(results) 
    for index in range(0, count):
        sheet.cell(row = index + 1, column = 1).value = input[index]
        sheet.cell(row = index + 1, column = 2).value = results[index]["lat"]
        sheet.cell(row = index + 1, column = 3).value = results[index]["lng"]
    filename = str(uuid.uuid4()) + ".xlsx"
    filepath = MEDIA_ROOT + "/" + filename
    book.save(filepath)
    return {'success': True, 'download_link': MEDIA_URL + filename}
